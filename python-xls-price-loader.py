# ===========================================================
# CONFIGURATION — edit only this section
# ===========================================================

WIN_OUTPUT_DIR   = r'C:\PATH\TO\OUTPUT'
WIN_TRACKER_FILE = r'C:\PATH\TO\TRACKER\pac.xlsx'
WIN_CSV_FILE     = r'C:\PATH\TO\OUTPUT\prezzi.csv'
WIN_HTML_FILE    = r'C:\PATH\TO\OUTPUT\portfolio-tracker.html'

LIN_OUTPUT_DIR   = '/path/to/output'
LIN_TRACKER_FILE = '/path/to/tracker/pac.xlsx'
LIN_CSV_FILE     = '/path/to/output/prezzi.csv'
LIN_HTML_FILE    = '/path/to/output/portfolio-tracker.html'

# Excel sheet name
SHEET_NAME = 'sheet name'

# ===========================================================

import os
import json
import platform
import yfinance as yf
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font

PLATFORM = platform.system()
print(f"Detected platform: {PLATFORM}")

if PLATFORM == 'Windows':
    OUTPUT_DIR   = WIN_OUTPUT_DIR
    TRACKER_FILE = WIN_TRACKER_FILE
    CSV_FILE     = WIN_CSV_FILE
    HTML_FILE    = WIN_HTML_FILE
    print("→ WINDOWS OFFICE mode")
elif PLATFORM == 'Linux':
    OUTPUT_DIR   = LIN_OUTPUT_DIR
    TRACKER_FILE = LIN_TRACKER_FILE
    CSV_FILE     = LIN_CSV_FILE
    HTML_FILE    = LIN_HTML_FILE
    print("→ LINUX LAPTOP mode")
else:
    raise Exception(f"Unsupported platform: {PLATFORM}")

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.chdir(OUTPUT_DIR)
print(f"Output dir: {OUTPUT_DIR} | CSV: {CSV_FILE} | Excel: {TRACKER_FILE} | HTML: {HTML_FILE}")

assets = {
    'GBP_EUR': 'GBPEUR=X',
    'USD_EUR': 'EURUSD=X',
    'XAU_USD': 'GC=F',
    'GOLD_INVESCO': 'SGLD.MI',
    'GLOBALE': 'SWDA.MI',
    'EMERG': 'EIMI.MI',
    'SMALL_CAP': 'IUSN.DE',
    'ENERGY': 'XDW0.MI',
    'R.EST': 'EPRA.MI',
    'XDWS': 'XDWS.MI',
    'ENEL': 'ENEL.MI',
    'BTC_SPOT_IBIT': 'IB1T.DE'
}

asset_meta = {
    'GBP_EUR': {'label': 'GBP/EUR', 'cat': 'FX', 'currency': 'EUR', 'decimals': 4},
    'USD_EUR': {'label': 'EUR/USD', 'cat': 'FX', 'currency': 'USD', 'decimals': 4},
    'XAU_USD': {'label': 'Gold Futures', 'cat': 'GOLD', 'currency': 'USD', 'decimals': 2},
    'GOLD_INVESCO': {'label': 'Invesco Physical Gold', 'cat': 'GOLD', 'currency': 'EUR', 'decimals': 2},
    'GLOBALE': {'label': 'iShares Core MSCI World', 'cat': 'ETF', 'currency': 'EUR', 'decimals': 2},
    'EMERG': {'label': 'iShares Core MSCI EM IMI', 'cat': 'ETF', 'currency': 'EUR', 'decimals': 3},
    'SMALL_CAP': {'label': 'iShares MSCI World Small Cap', 'cat': 'ETF', 'currency': 'EUR', 'decimals': 3},
    'ENERGY': {'label': 'Xtrackers MSCI World Energy', 'cat': 'ETF', 'currency': 'EUR', 'decimals': 2},
    'R.EST': {'label': 'iShares Developed Markets Property Yield', 'cat': 'ETF', 'currency': 'EUR', 'decimals': 2},
    'XDWS': {'label': 'Xtrackers MSCI World ESG', 'cat': 'ETF', 'currency': 'EUR', 'decimals': 2},
    'ENEL': {'label': 'Enel S.p.A.', 'cat': 'STOCK', 'currency': 'EUR', 'decimals': 3},
    'BTC_SPOT_IBIT': {'label': 'Bitcoin ETP', 'cat': 'CRYPTO', 'currency': 'EUR', 'decimals': 3}
}

celle_target = ['P2', 'P3', 'P6', 'P7', 'P9', 'P10', 'P11', 'P12',
                'P13', 'P14', 'P15', 'P16']

def safe_float(val):
    return float(val) if str(val) != 'N/A' and str(val) != 'Error' else 0

def get_live_price(t):
    price = None
    try:
        fi = t.fast_info
        price = fi.get('lastPrice', None)
    except:
        pass
    if price is None:
        try:
            info = t.info
            price = info.get('regularMarketPrice', info.get('currentPrice', None))
        except:
            pass
    if price is None:
        try:
            data_price = t.history(period='2d')
            if not data_price.empty and 'Close' in data_price.columns:
                price = data_price['Close'].iloc[-1]
        except:
            pass
    return round(float(price), 4) if price is not None else 'N/A'

def generate_dashboard_html(data_complete, generated_at):
    assets_js = []
    for key, ticker in assets.items():
        meta = asset_meta[key]
        assets_js.append({
            'id': key,
            'ticker': ticker,
            'label': meta['label'],
            'cat': meta['cat'],
            'currency': meta['currency'],
            'decimals': meta['decimals']
        })

    payload = {}
    for key, row in data_complete.items():
        payload[key] = {
            'Price': safe_float(row.get('Price')),
            '1d%': safe_float(row.get('1d%')),
            '1w%': safe_float(row.get('1w%')),
            '1m%': safe_float(row.get('1m%')),
            '6m%': safe_float(row.get('6m%')),
            '1y%': safe_float(row.get('1y%')),
            'ATH_Price': row.get('ATH_Price'),
            'ATH_Date': row.get('ATH_Date')
        }

    assets_json = json.dumps(assets_js, ensure_ascii=False)
    payload_json = json.dumps(payload, ensure_ascii=False)
    generated_date = generated_at.strftime('%d/%m/%Y')
    generated_time = generated_at.strftime('%H:%M:%S')

    html_template = r'''<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Portfolio Tracker</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:opsz,wght@9..40,300..700&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {
  --font-body: 'DM Sans', sans-serif;
  --font-mono: 'DM Mono', monospace;
  --text-xs: clamp(.75rem,.7rem + .25vw,.875rem);
  --text-sm: clamp(.875rem,.8rem + .35vw,1rem);
  --text-base: clamp(1rem,.95rem + .25vw,1.125rem);
  --text-lg: clamp(1.125rem,1rem + .75vw,1.5rem);
  --text-xl: clamp(1.5rem,1.2rem + 1.25vw,2.25rem);
  --radius-sm:.375rem; --radius-md:.5rem; --radius-lg:.75rem; --radius-xl:1rem; --radius-full:9999px;
  --transition:180ms cubic-bezier(.16,1,.3,1);
}
:root,[data-theme="light"]{
  --bg:#f7f6f2; --surface:#f9f8f5; --surface2:#f0ede8; --border:#d4d1ca; --divider:#dfdbd5;
  --text:#28251d; --muted:#7a7974; --faint:#afaea8; --primary:#01696f; --primary-hover:#0c4e54;
  --positive:#437a22; --negative:#a13544; --neutral:#7a7974; --gold:#c08800;
  --shadow-sm:0 1px 2px oklch(0.2 0.01 80 / .06); --shadow-md:0 4px 12px oklch(0.2 0.01 80 / .08);
}
[data-theme="dark"]{
  --bg:#131211; --surface:#1a1917; --surface2:#252320; --border:#353330; --divider:#2a2926;
  --text:#d4d2ce; --muted:#7a7876; --faint:#565450; --primary:#4f98a3; --primary-hover:#227f8b;
  --positive:#6daa45; --negative:#dd6974; --neutral:#7a7876; --gold:#e8af34;
  --shadow-sm:0 1px 2px oklch(0 0 0 / .25); --shadow-md:0 4px 12px oklch(0 0 0 / .35);
}
*{box-sizing:border-box;margin:0;padding:0}
body{min-height:100dvh;font-family:var(--font-body);background:var(--bg);color:var(--text);line-height:1.5}
button,input{font:inherit}
button{border:none;cursor:pointer}
.app{max-width:1400px;margin:0 auto;padding:1rem}
.topbar{display:flex;justify-content:space-between;align-items:center;gap:1rem;flex-wrap:wrap;background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:1rem 1.25rem;box-shadow:var(--shadow-sm);margin-bottom:1rem}
.brand{display:flex;align-items:center;gap:.75rem;font-weight:700}
.sub{font-size:var(--text-xs);color:var(--muted)}
.actions{display:flex;gap:.6rem;align-items:center;flex-wrap:wrap}
.btn{padding:.7rem 1rem;border-radius:var(--radius-md);background:var(--surface2);color:var(--text);border:1px solid var(--border);transition:all var(--transition);display:inline-flex;align-items:center;gap:.45rem}
.btn:hover{background:color-mix(in oklch,var(--primary) 8%,var(--surface2))}
.icon-btn{min-width:44px;justify-content:center;padding:.7rem .85rem}
.notice{background:color-mix(in oklch,var(--gold) 8%,var(--surface));border:1px solid color-mix(in oklch,var(--gold) 30%,var(--border));border-radius:var(--radius-md);padding:.85rem 1rem;color:var(--muted);font-size:var(--text-sm);margin-bottom:1rem}
.label{font-size:var(--text-xs);font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:var(--muted);margin:1.25rem 0 .6rem}
.ath{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:1rem 1.25rem;display:flex;gap:1.5rem;flex-wrap:wrap;align-items:center;box-shadow:var(--shadow-sm);margin-bottom:1rem}
.ath small{display:block;font-size:var(--text-xs);text-transform:uppercase;letter-spacing:.05em;color:var(--muted);margin-bottom:.2rem}
.ath strong{font-family:var(--font-mono);font-size:var(--text-lg)}
.ath-divider{width:1px;height:44px;background:var(--divider)}
.filters{display:flex;gap:.5rem;flex-wrap:wrap;margin-bottom:1rem}
.pill{padding:.35rem .75rem;border-radius:var(--radius-full);background:var(--surface);border:1px solid var(--border);color:var(--muted);font-size:var(--text-xs);font-weight:700}
.pill.active{background:var(--primary);border-color:var(--primary);color:#fff}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(min(220px,100%),1fr));gap:.75rem}
.card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:1rem;box-shadow:var(--shadow-sm)}
.card-top{display:flex;justify-content:space-between;align-items:flex-start;gap:.75rem;margin-bottom:.45rem}
.card-label{font-size:var(--text-xs);font-weight:700;letter-spacing:.05em;text-transform:uppercase;color:var(--muted)}
.card-price{font-family:var(--font-mono);font-size:var(--text-lg)}
.card-ticker{font-size:var(--text-xs);color:var(--faint);margin-top:.2rem}
.badge{padding:.15rem .45rem;border-radius:var(--radius-full);font-size:var(--text-xs);font-weight:700}
.pos{color:var(--positive)} .neg{color:var(--negative)} .neu{color:var(--neutral)}
.badge.pos{background:color-mix(in oklch,var(--positive) 15%,transparent)}
.badge.neg{background:color-mix(in oklch,var(--negative) 15%,transparent)}
.badge.neu{background:color-mix(in oklch,var(--neutral) 15%,transparent)}
.table-wrap{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);overflow:auto;box-shadow:var(--shadow-sm)}
table{width:100%;border-collapse:collapse;font-size:var(--text-xs)}
th,td{padding:.85rem .9rem;white-space:nowrap}
th{text-align:left;background:var(--surface2);color:var(--muted);font-weight:700;text-transform:uppercase;letter-spacing:.05em;border-bottom:1px solid var(--border)}
td{border-bottom:1px solid var(--divider)}
tbody tr:last-child td{border-bottom:none}
.num{text-align:right;font-family:var(--font-mono)}
.charts{display:grid;grid-template-columns:1fr 1fr;gap:1rem;margin-top:1rem}
.chart-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:1rem;box-shadow:var(--shadow-sm)}
.chart-title{font-size:var(--text-xs);font-weight:700;letter-spacing:.05em;text-transform:uppercase;color:var(--muted);margin-bottom:.5rem}
.chart-box{position:relative;height:240px}
.footer{margin-top:1rem;font-size:var(--text-xs);color:var(--faint);display:flex;justify-content:space-between;gap:1rem;flex-wrap:wrap}
@media(max-width:800px){.charts{grid-template-columns:1fr}.ath-divider{display:none}}
</style>
</head>
<body>
<div class="app">
  <div class="topbar">
    <div class="brand">
      <svg width="28" height="28" viewBox="0 0 28 28" fill="none" aria-hidden="true">
        <rect width="28" height="28" rx="7" fill="var(--primary)"/>
        <path d="M6 20L11 13L15 16L19 9L22 12" stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
        <circle cx="22" cy="12" r="2" fill="white"/>
      </svg>
      <div>
        <div>Portfolio Tracker</div>
        <div class="sub" id="updateInfo">Updated: __GENERATED_DATE__ · __GENERATED_TIME__</div>
      </div>
    </div>
    <div class="actions">
      <input id="csvFile" type="file" accept=".csv" hidden>
      <button class="btn" id="loadCsvBtn">Load CSV File</button>
      <button class="btn icon-btn" id="themeBtn" aria-label="Toggle theme" title="Toggle theme">
        <span id="themeIcon">☾</span>
      </button>
    </div>
  </div>

  <div class="notice">
    Standalone HTML dashboard generated automatically by your Python script. You can also overwrite the displayed data by manually loading an updated CSV file.
  </div>

  <div class="ath" id="athBanner" style="display:none">
    <div>
      <small>SWDA All-Time High</small>
      <strong id="athPrice">—</strong>
      <div class="sub" id="athDate">—</div>
    </div>
    <div class="ath-divider"></div>
    <div>
      <small>Current Price</small>
      <strong id="swdaNow">—</strong>
      <div class="sub">SWDA.MI</div>
    </div>
    <div class="ath-divider"></div>
    <div>
      <small>Distance from ATH</small>
      <strong id="athGap">—</strong>
      <div class="sub" id="athGapAbs">—</div>
    </div>
  </div>

  <div class="filters" id="filters">
    <button class="pill active" data-cat="ALL">All</button>
    <button class="pill" data-cat="FX">Currencies</button>
    <button class="pill" data-cat="GOLD">Gold</button>
    <button class="pill" data-cat="ETF">ETFs</button>
    <button class="pill" data-cat="STOCK">Stocks</button>
    <button class="pill" data-cat="CRYPTO">Crypto</button>
  </div>

  <div class="label">Prices</div>
  <div class="grid" id="cards"></div>

  <div class="label">Performance</div>
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>Asset</th>
          <th>Ticker</th>
          <th class="num">Price</th>
          <th class="num">1G</th>
          <th class="num">1S</th>
          <th class="num">1M</th>
          <th class="num">6M</th>
          <th class="num">1A</th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>

  <div class="label">Charts</div>
  <div class="charts">
    <div class="chart-card">
      <div class="chart-title">1-Week Performance</div>
      <div class="chart-box"><canvas id="chart1w"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title">1-Month Performance</div>
      <div class="chart-box"><canvas id="chart1m"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title">6-Month Performance</div>
      <div class="chart-box"><canvas id="chart6m"></canvas></div>
    </div>
    <div class="chart-card">
      <div class="chart-title">1-Year Performance</div>
      <div class="chart-box"><canvas id="chart1y"></canvas></div>
    </div>
  </div>

  <div class="footer">
    <span>Standalone file generated by PAC_auto_update.py v2.0.0</span>
    <span>No local dependencies required for viewing</span>
  </div>
</div>

<script>
const ASSETS = __ASSETS_JSON__;
const DEFAULT_DATA = __PAYLOAD_JSON__;
let DATA = JSON.parse(JSON.stringify(DEFAULT_DATA));
let activeCat = 'ALL';
let chart1w = null;
let chart1m = null;
let chart6m = null;
let chart1y = null;

function fmtPrice(v, d=2) {
  if (v === undefined || v === null || Number.isNaN(Number(v))) return '—';
  return Number(v).toLocaleString('en-US', { minimumFractionDigits: d, maximumFractionDigits: d });
}

function fmtPct(v) {
  if (v === undefined || v === null || Number.isNaN(Number(v))) return '—';
  const n = Number(v);
  return `${n >= 0 ? '+' : ''}${n.toFixed(2)}%`;
}

function cls(v) {
  const n = Number(v);
  if (Number.isNaN(n)) return 'neu';
  if (n > 0) return 'pos';
  if (n < 0) return 'neg';
  return 'neu';
}

function renderCards() {
  const wrap = document.getElementById('cards');
  wrap.innerHTML = '';
  ASSETS.filter(a => activeCat === 'ALL' || a.cat === activeCat).forEach(asset => {
    const row = DATA[asset.id] || {};
    const div = document.createElement('div');
    div.className = 'card';
    div.innerHTML = `
      <div class="card-top">
        <div class="card-label">${asset.label}</div>
        <div class="badge ${cls(row['1d%'])}">${fmtPct(row['1d%'])}</div>
      </div>
      <div class="card-price">${fmtPrice(row.Price, asset.decimals)} <span style="font-size:.72em;color:var(--muted)">${asset.currency}</span></div>
      <div class="card-ticker">${asset.ticker}</div>
    `;
    wrap.appendChild(div);
  });
}

function renderTable() {
  const tbody = document.getElementById('tbody');
  tbody.innerHTML = '';
  ASSETS.filter(a => activeCat === 'ALL' || a.cat === activeCat).forEach(asset => {
    const row = DATA[asset.id] || {};
    const tr = document.createElement('tr');
    tr.innerHTML = `
      <td><strong>${asset.label}</strong></td>
      <td>${asset.ticker}</td>
      <td class="num">${fmtPrice(row.Price, asset.decimals)}</td>
      <td class="num ${cls(row['1d%'])}">${fmtPct(row['1d%'])}</td>
      <td class="num ${cls(row['1w%'])}">${fmtPct(row['1w%'])}</td>
      <td class="num ${cls(row['1m%'])}">${fmtPct(row['1m%'])}</td>
      <td class="num ${cls(row['6m%'])}">${fmtPct(row['6m%'])}</td>
      <td class="num ${cls(row['1y%'])}">${fmtPct(row['1y%'])}</td>
    `;
    tbody.appendChild(tr);
  });
}

function renderAth() {
  const swda = DATA['GLOBALE'];
  const banner = document.getElementById('athBanner');
  if (!swda || swda.ATH_Price === null || swda.ATH_Price === undefined || Number.isNaN(Number(swda.ATH_Price))) {
    banner.style.display = 'none';
    return;
  }
  banner.style.display = 'flex';
  document.getElementById('athPrice').textContent = fmtPrice(swda.ATH_Price, 2) + ' €';
  document.getElementById('athDate').textContent = swda.ATH_Date || '—';
  document.getElementById('swdaNow').textContent = fmtPrice(swda.Price, 2) + ' €';
  const gap = ((Number(swda.Price) - Number(swda.ATH_Price)) / Number(swda.ATH_Price)) * 100;
  const gapEl = document.getElementById('athGap');
  gapEl.textContent = fmtPct(gap);
  gapEl.className = cls(gap);
  const abs = Number(swda.Price) - Number(swda.ATH_Price);
  document.getElementById('athGapAbs').textContent = `${abs >= 0 ? '+' : ''}${abs.toFixed(2)} € from ATH`;
}

function chartColors(arr) {
  return arr.map(v => Number(v) >= 0 ? 'rgba(109,170,69,.75)' : 'rgba(221,105,116,.75)');
}

function renderCharts() {
  const selected = ASSETS.filter(a => activeCat === 'ALL' || a.cat === activeCat);
  const labels = selected.map(a => a.label.length > 10 ? a.label.slice(0, 10) : a.label);
  const vals1w = selected.map(a => Number((DATA[a.id] || {})['1w%'] || 0));
  const vals1m = selected.map(a => Number((DATA[a.id] || {})['1m%'] || 0));
  const vals6m = selected.map(a => Number((DATA[a.id] || {})['6m%'] || 0));
  const vals1y = selected.map(a => Number((DATA[a.id] || {})['1y%'] || 0));
  const muted = getComputedStyle(document.documentElement).getPropertyValue('--muted').trim();
  const common = {
    responsive: true,
    maintainAspectRatio: false,
    plugins: { legend: { display: false } },
    scales: { x: { ticks: { color: muted } }, y: { ticks: { color: muted } } }
  };
  if (chart1w) chart1w.destroy();
  if (chart1m) chart1m.destroy();
  if (chart6m) chart6m.destroy();
  if (chart1y) chart1y.destroy();
  chart1w = new Chart(document.getElementById('chart1w'), {
    type: 'bar',
    data: { labels, datasets: [{ data: vals1w, backgroundColor: chartColors(vals1w), borderRadius: 4 }] },
    options: common
  });
  chart1m = new Chart(document.getElementById('chart1m'), {
    type: 'bar',
    data: { labels, datasets: [{ data: vals1m, backgroundColor: chartColors(vals1m), borderRadius: 4 }] },
    options: common
  });
  chart6m = new Chart(document.getElementById('chart6m'), {
    type: 'bar',
    data: { labels, datasets: [{ data: vals6m, backgroundColor: chartColors(vals6m), borderRadius: 4 }] },
    options: common
  });
  chart1y = new Chart(document.getElementById('chart1y'), {
    type: 'bar',
    data: { labels, datasets: [{ data: vals1y, backgroundColor: chartColors(vals1y), borderRadius: 4 }] },
    options: common
  });
}

function renderAll() {
  renderCards();
  renderTable();
  renderAth();
  renderCharts();
}

function splitCsvLine(line) {
  const result = [];
  let current = '';
  let inQuotes = false;
  for (let i = 0; i < line.length; i++) {
    const ch = line[i];
    if (ch === '"') {
      if (inQuotes && line[i + 1] === '"') {
        current += '"';
        i++;
      } else {
        inQuotes = !inQuotes;
      }
    } else if (ch === ',' && !inQuotes) {
      result.push(current);
      current = '';
    } else {
      current += ch;
    }
  }
  result.push(current);
  return result;
}

function parseCsv(text) {
  const lines = text.trim().split(/\\r?\\n/).filter(Boolean);
  if (lines.length < 2) throw new Error('Empty or invalid CSV');
  const headers = splitCsvLine(lines[0]);
  const parsed = {};
  for (let i = 1; i < lines.length; i++) {
    const cols = splitCsvLine(lines[i]);
    const key = cols[0];
    if (!key) continue;
    const row = {};
    for (let j = 1; j < headers.length; j++) {
      const h = headers[j];
      const raw = cols[j] ?? '';
      if (raw === '') {
        row[h] = null;
      } else {
        const n = Number(raw);
        row[h] = Number.isNaN(n) ? raw : n;
      }
    }
    parsed[key] = row;
  }
  return parsed;
}

function loadCsvFile(file) {
  const reader = new FileReader();
  reader.onload = (e) => {
    try {
      DATA = parseCsv(e.target.result);
      document.getElementById('updateInfo').textContent = `CSV loaded: ${file.name}`;
      renderAll();
    } catch (err) {
      alert('Error reading CSV file: ' + err.message);
    }
  };
  reader.readAsText(file, 'utf-8');
}

document.getElementById('loadCsvBtn').addEventListener('click', () => document.getElementById('csvFile').click());
document.getElementById('csvFile').addEventListener('change', (e) => {
  const file = e.target.files?.[0];
  if (file) loadCsvFile(file);
});

document.getElementById('filters').addEventListener('click', (e) => {
  const btn = e.target.closest('.pill');
  if (!btn) return;
  document.querySelectorAll('.pill').forEach(p => p.classList.remove('active'));
  btn.classList.add('active');
  activeCat = btn.dataset.cat;
  renderAll();
});

document.getElementById('themeBtn').addEventListener('click', () => {
  const root = document.documentElement;
  const isDark = root.getAttribute('data-theme') === 'dark';
  root.setAttribute('data-theme', isDark ? 'light' : 'dark');
  document.getElementById('themeIcon').textContent = isDark ? '☀' : '☾';
  renderCharts();
});

document.getElementById('themeIcon').textContent = document.documentElement.getAttribute('data-theme') === 'dark' ? '☾' : '☀';
renderAll();
</script>
</body>
</html>'''

    html_template = html_template.replace('__ASSETS_JSON__', assets_json)
    html_template = html_template.replace('__PAYLOAD_JSON__', payload_json)
    html_template = html_template.replace('__GENERATED_DATE__', generated_date)
    html_template = html_template.replace('__GENERATED_TIME__', generated_time)
    return html_template

print(f"Updating live prices + SWDA ATH... {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

data_complete = {}
for name, ticker in assets.items():
    try:
        t = yf.Ticker(ticker)
        price = get_live_price(t)
        print(f"{name}: {price}")

        try:
            hist = t.history(period='1y')
            if not hist.empty:
                close = hist['Close']
                perf_1d = round(((close.iloc[-1] / close.iloc[-2] - 1) * 100), 2) if len(close) >= 2 else 0
                perf_1w = round(((close.iloc[-1] / close.iloc[-5] - 1) * 100), 2) if len(close) >= 5 else 0
                perf_1m = round(((close.iloc[-1] / close.iloc[-21] - 1) * 100), 2) if len(close) >= 21 else 0
                perf_6m = round(((close.iloc[-1] / close.iloc[-126] - 1) * 100), 2) if len(close) >= 126 else 0
                perf_1y = round(((close.iloc[-1] / close.iloc[0] - 1) * 100), 2)
            else:
                raise Exception("No 1y data")
        except:
            hist_mo = t.history(period='1mo')
            if not hist_mo.empty:
                close = hist_mo['Close']
                perf_1d = round(((close.iloc[-1] / close.iloc[-2] - 1) * 100), 2) if len(close) >= 2 else 0
                perf_1w = round(((close.iloc[-1] / close.iloc[-5] - 1) * 100), 2) if len(close) >= 5 else 0
                perf_1m = round(((close.iloc[-1] / close.iloc[0] - 1) * 100), 2)
                perf_6m = perf_1m
                perf_1y = perf_1m
            else:
                perf_1d = perf_1w = perf_1m = perf_6m = perf_1y = 0

        ath_price = None
        ath_date = None
        if name == 'GLOBALE':
            try:
                hist_max = t.history(period='max', auto_adjust=False)
                hist_today = t.history(period='1d', auto_adjust=False)
                candidates = []
                if not hist_max.empty and 'High' in hist_max.columns:
                    candidates.append((float(hist_max['High'].max()), hist_max['High'].idxmax().strftime('%d/%m/%Y')))
                if not hist_today.empty and 'High' in hist_today.columns:
                    candidates.append((float(hist_today['High'].max()), hist_today['High'].idxmax().strftime('%d/%m/%Y')))
                if candidates:
                    ath_price, ath_date = max(candidates, key=lambda x: x[0])
                    ath_price = round(ath_price, 2)
                    print(f"SWDA ATH: High {ath_price} on {ath_date} (historical rows: {len(hist_max)})")
            except Exception as ath_e:
                print(f"SWDA ATH error: {ath_e}")

        data_complete[name] = {
            'Price': price, '1d%': perf_1d, '1w%': perf_1w,
            '1m%': perf_1m, '6m%': perf_6m, '1y%': perf_1y,
            'ATH_Price': ath_price, 'ATH_Date': ath_date
        }

    except Exception as e:
        print(f"❌ Error {name} ({ticker}): {e}")
        data_complete[name] = {'Price': 'Error', '1d%': 0, '1w%': 0, '1m%': 0, '6m%': 0, '1y%': 0, 'ATH_Price': None, 'ATH_Date': None}

df = pd.DataFrame(data_complete).T
print(df.round(2))
df.to_csv(CSV_FILE, index=True)
print(f'\n✓ CSV saved: {CSV_FILE}')

try:
    print('\n--- Excel P/Q/R/S/T/U + V9 HighATH W9 Date + Green/Red ---')
    wb = openpyxl.load_workbook(TRACKER_FILE, read_only=False, keep_vba=True)
    ws = wb[SHEET_NAME]

    for i, asset in enumerate(assets.keys()):
        row_data = data_complete[asset]
        num_riga = celle_target[i][1:]

        ws[f'P{num_riga}'] = safe_float(row_data['Price'])

        valore_1d = safe_float(row_data['1d%']) / 100
        ws[f'Q{num_riga}'] = valore_1d
        ws[f'Q{num_riga}'].number_format = '0.00%'

        for col, perf_key in zip(['R', 'S', 'T', 'U'], ['1w%', '1m%', '6m%', '1y%']):
            valore = safe_float(row_data[perf_key]) / 100
            cella = f'{col}{num_riga}'
            ws[cella] = valore
            ws[cella].number_format = '0.00%'
            if valore > 0:
                ws[cella].font = Font(name='Bookman Old Style', color='00B050', bold=True, size=11)
            elif valore < 0:
                ws[cella].font = Font(name='Bookman Old Style', color='FF0000', bold=True, size=11)
            else:
                ws[cella].font = Font(name='Bookman Old Style', size=11)
            print(f'{asset} {col}{num_riga}: {valore:+.2%}')

        if valore_1d > 0:
            ws[f'Q{num_riga}'].font = Font(name='Bookman Old Style', color='00B050', bold=True, size=11)
        elif valore_1d < 0:
            ws[f'Q{num_riga}'].font = Font(name='Bookman Old Style', color='FF0000', bold=True, size=11)
        else:
            ws[f'Q{num_riga}'].font = Font(name='Bookman Old Style', size=11)

        if i == 4 and row_data.get('ATH_Price') is not None:
            ws['V9'] = safe_float(row_data['ATH_Price'])
            ws['V9'].number_format = '0.00'
            ws['W9'] = row_data.get('ATH_Date', '')
            print(f'{asset} V9 HighATH: {row_data["ATH_Price"]} - W9 Date: {row_data["ATH_Date"]}')

    update_time = datetime.now()
    ws['N19'] = update_time.strftime('%d/%m/%Y')
    ws['N20'] = update_time.strftime('%H:%M:%S')
    print('TIMESTAMP → N19/N20 ✓')

    intestazioni = {'P18': 'Prices', 'Q18': '%day', 'R18': '%wk', 'S18': '%1 Mth', 'T18': '%6 Mth', 'U18': '%1 YR'}
    for cella, testo in intestazioni.items():
        ws[cella] = testo
        ws[cella].font = Font(name='Bookman Old Style', bold=True, size=11)

    wb.save(TRACKER_FILE)
    print(f'\n✓ Excel saved: {TRACKER_FILE}')

except Exception as e:
    print(f'❌ Excel error: {e}')

try:
    generated_at = datetime.now()
    html_content = generate_dashboard_html(data_complete, generated_at)
    with open(HTML_FILE, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f'✓ HTML dashboard saved: {HTML_FILE}')
except Exception as e:
    print(f'❌ HTML error: {e}')

print(f'\n✅ Complete v2.0.0: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
print(f'PLATFORM: {PLATFORM} | Assets: {len(assets)} | CSV: {CSV_FILE} | HTML: {HTML_FILE}')